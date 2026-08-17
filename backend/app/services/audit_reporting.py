"""Auditor-only request ledger aggregation and native Excel export."""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import hashlib
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import select

from app.db import ApprovalStepModel, AuditLogModel, ChangeRequestModel, Repository, UserModel


def _iso(value: datetime | None) -> str | None:
    if not value:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).isoformat()


def _event_payload(row: AuditLogModel) -> dict[str, Any]:
    return {
        "id": row.event_id,
        "event": row.event_type,
        "actor": row.actor,
        "at": _iso(row.created_at),
        **(row.payload or {}),
    }


def verify_audit_chain(events: list[dict[str, Any]]) -> bool:
    previous = "GENESIS"
    for event in (item for item in events if item.get("current_hash")):
        current = event.get("current_hash")
        prior = event.get("prior_hash")
        # Older operational events were intentionally not hash-linked. When one
        # sits between governed workflow events, `_audit` starts a new segment
        # with an empty root. Each segment must still be internally contiguous.
        if prior != previous:
            if event.get("chain_version") or prior not in {"GENESIS", ""}:
                return False
        # Repository payload hashes do not include storage/display-only fields.
        stored_payload = {key: value for key, value in event.items() if key not in {"id", "actor", "actor_name"}}
        stored_payload.pop("current_hash", None)
        # Audit payloads carry their authoritative transition time as `at`.
        candidate = hashlib.sha256(json.dumps(stored_payload, sort_keys=True, default=str).encode()).hexdigest()
        if candidate != current:
            return False
        previous = current
    return True


def audit_event_rows(repo: Repository) -> list[dict[str, Any]]:
    with repo.session() as session:
        rows = session.scalars(select(AuditLogModel).order_by(AuditLogModel.id)).all()
        users = session.scalars(select(UserModel)).all()
    names = {item.user_id: item.display_name for item in users} | {item.email: item.display_name for item in users}
    return [{**_event_payload(row), "actor_name": names.get(row.actor, row.actor or "System")} for row in rows]


def build_audit_request_rows(
    repo: Repository,
    site_id: str | None = None,
    search: str | None = None,
    status: str | None = None,
    decision: str | None = None,
) -> dict[str, Any]:
    with repo.session() as session:
        requests = session.scalars(select(ChangeRequestModel).order_by(ChangeRequestModel.created_at.desc())).all()
        steps = session.scalars(select(ApprovalStepModel).order_by(ApprovalStepModel.request_id, ApprovalStepModel.order_index)).all()
        users = session.scalars(select(UserModel)).all()
    events = audit_event_rows(repo)
    names = {item.user_id: item.display_name for item in users} | {item.email: item.display_name for item in users}
    steps_by_request: dict[str, list[ApprovalStepModel]] = {}
    events_by_request: dict[str, list[dict[str, Any]]] = {}
    for step in steps:
        steps_by_request.setdefault(step.request_id, []).append(step)
    for event in events:
        if event.get("request_id"):
            events_by_request.setdefault(str(event["request_id"]), []).append(event)

    result: list[dict[str, Any]] = []
    for request in requests:
        request_steps = steps_by_request.get(request.request_id, [])
        request_events = events_by_request.get(request.request_id, [])
        decided_steps = [step for step in request_steps if step.decided_at]
        final_step = max(decided_steps, key=lambda step: step.decided_at) if decided_steps else None
        submitted_event = next((item for item in request_events if item.get("event") == "change_submitted"), None)
        verified_event = next((item for item in reversed(request_events) if item.get("event") == "change_verified"), None)
        latest_event = request_events[-1] if request_events else None
        decision = final_step.decision if final_step else None
        row = {
            "request_id": request.request_id,
            "title": (request.payload or {}).get("title") or request.anomaly_id,
            "site_id": request.site_id,
            "severity": request.severity,
            "impact_euros": request.impact_euros,
            "status": request.status,
            "decision": decision,
            "requester_id": request.requested_by,
            "requester_name": names.get(request.requested_by, request.requested_by),
            "requested_at": _iso(request.created_at),
            "submitted_at": submitted_event.get("at") if submitted_event else None,
            "final_approver_id": final_step.decided_by if final_step else None,
            "final_approver_name": names.get(final_step.decided_by, final_step.decided_by) if final_step and final_step.decided_by else None,
            "decided_at": _iso(final_step.decided_at) if final_step else None,
            "approved_at": _iso(final_step.decided_at) if final_step and decision == "approved" else None,
            "rejected_at": _iso(final_step.decided_at) if final_step and decision == "rejected" else None,
            "decision_comment": final_step.comment if final_step else None,
            "verified_at": verified_event.get("at") if verified_event else (request.payload or {}).get("verification", {}).get("verified_at"),
            "policy_version": request.policy_version,
            "source_hash": request.source_hash,
            "latest_audit_hash": latest_event.get("current_hash") if latest_event else None,
            "approval_stages": [{
                "stage": step.stage,
                "required_role": step.required_role,
                "assigned_to": step.assigned_to,
                "assigned_name": names.get(step.assigned_to, step.assigned_to) if step.assigned_to else None,
                "status": step.status,
                "decision": step.decision,
                "decided_by": step.decided_by,
                "decider_name": names.get(step.decided_by, step.decided_by) if step.decided_by else None,
                "decided_at": _iso(step.decided_at),
                "comment": step.comment,
            } for step in request_steps],
        }
        haystack = " ".join(str(row.get(key) or "") for key in ("request_id", "title", "site_id", "status", "decision", "requester_name", "final_approver_name")).lower()
        if site_id and site_id != "all" and request.site_id != site_id:
            continue
        if status and status != "all" and request.status != status:
            continue
        if decision and decision != "all" and row["decision"] != decision:
            continue
        if search and search.lower() not in haystack:
            continue
        result.append(row)
    return {"items": result, "total": len(result), "chain_verified": verify_audit_chain(events), "event_count": len(events)}


def _excel_datetime(value: str | None):
    if not value:
        return None
    parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if parsed.tzinfo:
        parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
    return parsed


def _excel_safe(value):
    """Prevent exported user text from becoming an executable Excel formula."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _style_sheet(sheet, table_name: str, date_columns: set[int], widths: list[int]) -> None:
    navy, green, neon, blue = "002733", "008C82", "C2FE06", "8CBEE6"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 28
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=neon, bold=True, size=10)
        cell.alignment = Alignment(vertical="center")
    border = Border(bottom=Side(style="thin", color=blue))
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.column in date_columns and cell.value:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else "A"].width = width
    if sheet.max_row >= 2:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.auto_filter.ref = sheet.dimensions


def build_audit_workbook(request_rows: list[dict[str, Any]], events: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    requests_sheet = workbook.active
    requests_sheet.title = "Request Decisions"
    request_headers = [
        "Request ID", "Title", "Site", "Severity", "Impact (EUR)", "Status", "Decision",
        "Requester", "Requester ID", "Requested At (UTC)", "Submitted At (UTC)",
        "Final Approver", "Approver ID", "Decided At (UTC)", "Approved At (UTC)",
        "Rejected At (UTC)", "Decision Comment", "Verified At (UTC)", "Policy Version",
        "Approval Route", "Source Hash", "Latest Audit Hash",
    ]
    requests_sheet.append(request_headers)
    for row in request_rows:
        route = " | ".join(
            f"{stage['required_role']}: {stage.get('assigned_name') or 'unassigned'} → {stage.get('decision') or stage['status']}"
            for stage in row.get("approval_stages", [])
        )
        requests_sheet.append([_excel_safe(value) for value in [
            row["request_id"], row["title"], row["site_id"], row["severity"], row["impact_euros"], row["status"], row.get("decision"),
            row["requester_name"], row["requester_id"], _excel_datetime(row.get("requested_at")), _excel_datetime(row.get("submitted_at")),
            row.get("final_approver_name"), row.get("final_approver_id"), _excel_datetime(row.get("decided_at")), _excel_datetime(row.get("approved_at")),
            _excel_datetime(row.get("rejected_at")), row.get("decision_comment"), _excel_datetime(row.get("verified_at")), row["policy_version"], route,
            row["source_hash"], row.get("latest_audit_hash"),
        ]])
    for cell in requests_sheet["E"][1:]:
        cell.number_format = '€#,##0'
    _style_sheet(requests_sheet, "RequestDecisionTable", {10, 11, 14, 15, 16, 18}, [20, 34, 14, 12, 16, 18, 14, 24, 18, 21, 21, 24, 18, 21, 21, 21, 34, 21, 14, 48, 22, 22])

    event_sheet = workbook.create_sheet("Immutable Event Log")
    event_headers = ["Event ID", "Timestamp (UTC)", "Event", "Actor", "Role", "Site", "Request ID", "Policy Version", "Prior Hash", "Current Hash"]
    event_sheet.append(event_headers)
    for event in events:
        event_sheet.append([_excel_safe(value) for value in [
            event.get("id"), _excel_datetime(event.get("at")), event.get("event"), event.get("actor_name") or event.get("actor"),
            event.get("role"), event.get("site_id"), event.get("request_id"), event.get("policy_version"), event.get("prior_hash"), event.get("current_hash"),
        ]])
    _style_sheet(event_sheet, "ImmutableEventTable", {2}, [38, 21, 28, 24, 18, 14, 20, 14, 22, 22])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
