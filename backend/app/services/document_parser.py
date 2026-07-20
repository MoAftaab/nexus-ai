"""Parse, preview, cross-check and index operational documents."""
from __future__ import annotations

import base64
import csv
from io import BytesIO, StringIO
from pathlib import Path
import uuid

from app.config import Settings
from app.models import DocumentInspection
from app.services.knowledge_base import index_document


def extract_text(filename: str, raw: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in {".txt", ".md", ".csv", ".json", ".xml"}:
        decoded = raw.decode("utf-8", errors="ignore")
        if suffix == ".csv":
            return "\n".join(" | ".join(row) for row in list(csv.reader(StringIO(decoded)))[:500])
        return decoded
    if suffix == ".pdf":
        from pypdf import PdfReader
        return "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(raw)).pages)
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        rows = []
        for worksheet in workbook.worksheets:
            rows.append(f"## Sheet: {worksheet.title}")
            rows.extend(" | ".join("" if value is None else str(value) for value in row) for row in worksheet.iter_rows(values_only=True, max_row=500))
        return "\n".join(rows)
    return raw.decode("utf-8", errors="ignore")


def render_preview(filename: str, raw: bytes, preview_path: Path) -> Path | None:
    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".pdf":
            import fitz
            document = fitz.open(stream=raw, filetype="pdf")
            page = document.load_page(0)
            pixmap = page.get_pixmap(matrix=fitz.Matrix(1.4, 1.4), alpha=False)
            pixmap.save(str(preview_path))
            return preview_path
        if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
            from PIL import Image
            image = Image.open(BytesIO(raw)).convert("RGB")
            image.thumbnail((1600, 1600)); image.save(preview_path, "PNG")
            return preview_path
    except Exception:
        return None
    return None


async def extract_with_openai_vision(settings: Settings, filename: str, preview: Path | None, text: str) -> str:
    """Use the mandated mini model only when a scanned document has no readable text."""
    if text.strip() or not preview or not settings.openai_api_key:
        return text
    from openai import AsyncOpenAI
    encoded = base64.b64encode(preview.read_bytes()).decode("ascii")
    client = AsyncOpenAI(api_key=settings.openai_api_key)
    response = await client.responses.create(model=settings.openai_model, instructions="Extract operational identifiers, quantities, dates, PPAP/VDA references and any release conditions from this logistics document. Return concise plain text only.", input=[{"role": "user", "content": [{"type": "input_text", "text": f"Extract this scanned document named {filename}."}, {"type": "input_image", "image_url": f"data:image/png;base64,{encoded}"}]}])
    return response.output_text


async def inspect_and_index(store, settings: Settings, filename: str, raw: bytes) -> DocumentInspection:
    document_id = f"DOC-UP-{uuid.uuid4().hex[:12].upper()}"
    uploads = Path(settings.uploads_dir).resolve(); uploads.mkdir(parents=True, exist_ok=True)
    suffix = Path(filename).suffix.lower() or ".bin"
    stored = uploads / f"{document_id}{suffix}"
    stored.write_bytes(raw)
    preview = render_preview(filename, raw, uploads / f"{document_id}.preview.png")
    text = await extract_with_openai_vision(settings, filename, preview, extract_text(filename, raw))
    result = store.inspect_document(filename, text)
    markdown_record = index_document(filename, result.type, text, result.summary, result.mismatches)
    markdown_path = str((Path(__file__).resolve().parents[2] / "knowledge" / "ingested" / markdown_record).resolve())
    result.fields.append({"label": "Knowledge record", "value": markdown_record})
    store.record_document(document_id, result, str(stored), markdown_path)
    result.document_id = document_id
    result.preview_url = f"/api/documents/{document_id}/preview" if preview else None
    return result
