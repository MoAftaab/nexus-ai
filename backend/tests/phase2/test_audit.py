from __future__ import annotations

from app.db import AuditLogModel, Repository
from app.config import get_settings
from io import BytesIO
from openpyxl import load_workbook
from app.services.audit_reporting import build_audit_workbook


def test_approval_decision_creates_hash_linked_append_only_audit(client, login_as):
    requester = login_as(client, "operator1@nexusai.demo")
    findings = [item for item in client.get("/api/anomalies").json()["items"] if item["actions"]]
    finding = next((item for item in findings if item["severity"] == "low" and item["impact"] < 25_000), findings[0])
    preview = client.post("/api/changes/preview", headers=requester, json={"anomaly_id": finding["id"], "action_id": finding["actions"][0]["id"]}).json()
    request = client.post("/api/changes", headers=requester, json=preview).json()
    client.post(f"/api/changes/{request['request_id']}/submit", headers=requester)
    role = client.get(f"/api/changes/{request['request_id']}", headers=requester).json()["active_step"]["required_role"]
    approver = login_as(client, {"lead": "lead1@nexusai.demo", "manager": "manager1@nexusai.demo", "quality_compliance": "quality1@nexusai.demo", "director": "director@nexusai.demo"}[role])
    decision = client.post(f"/api/changes/{request['request_id']}/return", headers=approver, json={"comment": "needs evidence"})
    assert decision.status_code == 200
    audit = client.get("/api/audit", headers=login_as(client, "auditor@nexusai.demo")).json()["items"]
    change_events = [row for row in audit if row.get("request_id") == request["request_id"]]
    assert change_events
    assert all(row.get("role") and row.get("site_id") and row.get("at") for row in change_events)
    assert all("prior_hash" in row and "current_hash" in row for row in change_events)
    assert all("snapshot_hash" in row for row in change_events)


def test_demo_reset_preserves_audit_history(client, login_as):
    headers = login_as(client, "auditor@nexusai.demo")
    before = client.get("/api/audit", headers=headers).json()["items"]
    assert client.post("/api/demo/reset").status_code == 200
    after = client.get("/api/audit", headers=headers).json()["items"]
    assert len(after) >= len(before)


def test_audit_chain_has_no_duplicate_event_ids():
    repository = Repository(get_settings())
    with repository.session() as session:
        rows = session.query(AuditLogModel).all()
        assert len({row.event_id for row in rows}) == len(rows)


def test_audit_archive_is_available_only_to_the_auditor_role(client, login_as):
    assert client.get("/api/audit").status_code == 401
    for email in ("operator1@nexusai.demo", "lead1@nexusai.demo", "director@nexusai.demo", "admin@nexusai.demo"):
        assert client.get("/api/audit", headers=login_as(client, email)).status_code == 403

    response = client.get("/api/audit", headers=login_as(client, "auditor@nexusai.demo"))
    assert response.status_code == 200
    assert "items" in response.json()


def test_auditor_receives_named_timestamped_approval_history_and_other_roles_do_not(client, login_as):
    requester = login_as(client, "operator1@nexusai.demo")
    finding = next(
        item for item in client.get("/api/anomalies").json()["items"]
        if item["actions"] and item["severity"] == "low" and item["impact"] < 25_000
    )
    preview = client.post(
        "/api/changes/preview",
        headers=requester,
        json={"anomaly_id": finding["id"], "action_id": finding["actions"][0]["id"]},
    ).json()
    request = client.post("/api/changes", headers=requester, json=preview).json()
    client.post(f"/api/changes/{request['request_id']}/submit", headers=requester)
    approved = client.post(
        f"/api/changes/{request['request_id']}/approve",
        headers=login_as(client, "lead1@nexusai.demo"),
        json={"comment": "Evidence checked"},
    )
    assert approved.status_code == 200, approved.text

    operator_detail = client.get(f"/api/changes/{request['request_id']}", headers=requester).json()
    assert "audit_history" not in operator_detail
    assert "decided_by" not in operator_detail["steps"][0]
    assert "decided_at" not in operator_detail["steps"][0]
    assert "comment" not in operator_detail["steps"][0]

    auditor = login_as(client, "auditor@nexusai.demo")
    auditor_detail = client.get(f"/api/changes/{request['request_id']}", headers=auditor).json()
    history = auditor_detail["audit_history"]
    assert history["requested_by"]["display_name"] == "Operator Wolfsburg"
    assert history["requested_at"]
    assert history["submitted_at"]
    assert history["approvals"][0]["approver"]["display_name"] == "Lead Wolfsburg"
    assert history["approvals"][0]["decision"] == "approved"
    assert history["approvals"][0]["decided_at"]
    assert history["approvals"][0]["comment"] == "Evidence checked"

    archive = client.get("/api/audit", headers=auditor).json()["items"]
    approval_event = next(
        item for item in archive
        if item.get("request_id") == request["request_id"] and item["event"] == "approval_decided"
    )
    assert approval_event["actor_name"] == "Lead Wolfsburg"


def test_audit_request_table_has_request_and_decision_times_for_rejections(client, login_as):
    requester = login_as(client, "operator1@nexusai.demo")
    finding = next(item for item in client.get("/api/anomalies").json()["items"] if item["actions"] and item["severity"] == "low" and item["impact"] < 25_000)
    preview = client.post("/api/changes/preview", headers=requester, json={"anomaly_id": finding["id"], "action_id": finding["actions"][0]["id"]}).json()
    request = client.post("/api/changes", headers=requester, json=preview).json()
    client.post(f"/api/changes/{request['request_id']}/submit", headers=requester)
    rejected = client.post(f"/api/changes/{request['request_id']}/reject", headers=login_as(client, "lead1@nexusai.demo"), json={"comment": "Evidence does not reconcile"})
    assert rejected.status_code == 200

    auditor = login_as(client, "auditor@nexusai.demo")
    response = client.get("/api/audit/requests", headers=auditor)
    assert response.status_code == 200
    row = next(item for item in response.json()["items"] if item["request_id"] == request["request_id"])
    assert row["requested_at"]
    assert row["submitted_at"]
    assert row["decision"] == "rejected"
    assert row["decided_at"]
    assert row["requester_name"] == "Operator Wolfsburg"
    assert row["final_approver_name"] == "Lead Wolfsburg"
    assert row["decision_comment"] == "Evidence does not reconcile"
    assert row["approval_stages"][0]["decided_at"]
    assert response.json()["chain_verified"] is True


def test_auditor_can_export_real_excel_workbook_and_other_roles_cannot(client, login_as):
    auditor = login_as(client, "auditor@nexusai.demo")
    response = client.get("/api/audit/export.xlsx", headers=auditor)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "warehouse-control-tower-audit" in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), read_only=False, data_only=False)
    assert workbook.sheetnames == ["Request Decisions", "Immutable Event Log"]
    request_headers = [cell.value for cell in workbook["Request Decisions"][1]]
    assert "Requested At (UTC)" in request_headers
    assert "Decided At (UTC)" in request_headers
    assert "Final Approver" in request_headers
    assert workbook["Request Decisions"].auto_filter.ref
    assert workbook["Immutable Event Log"].auto_filter.ref
    assert client.get("/api/audit/export.xlsx", headers=login_as(client, "manager1@nexusai.demo")).status_code == 403


def test_excel_export_neutralizes_formula_injection():
    rows = [{
        "request_id": "CR-SAFE", "title": "=HYPERLINK(\"https://invalid\")", "site_id": "wolfsburg", "severity": "low",
        "impact_euros": 10, "status": "rejected", "decision": "rejected", "requester_name": "Operator", "requester_id": "operator1",
        "requested_at": None, "submitted_at": None, "final_approver_name": "Lead", "final_approver_id": "lead1", "decided_at": None,
        "approved_at": None, "rejected_at": None, "decision_comment": "+cmd", "verified_at": None, "policy_version": 1,
        "approval_stages": [], "source_hash": "source", "latest_audit_hash": "audit",
    }]
    workbook = load_workbook(BytesIO(build_audit_workbook(rows, [])), data_only=False)
    assert workbook["Request Decisions"]["B2"].value.startswith("'=")
    assert workbook["Request Decisions"]["Q2"].value == "'+cmd"
